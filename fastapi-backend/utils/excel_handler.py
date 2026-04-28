from io import BytesIO
import openpyxl

def read_excel(file_stream):
    wb = openpyxl.load_workbook(file_stream)
    ws = wb.active
    data = []
    
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    
    return data

def write_excel(headers, data):
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws.append(headers)
    
    for row in data:
        ws.append(row)
    
    wb.save(output)
    output.seek(0)
    return output

def generate_single_column_template(header):
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([header])
    wb.save(output)
    output.seek(0)
    return output

def generate_double_column_template(header1, header2):
    output = BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([header1, header2])
    wb.save(output)
    output.seek(0)
    return output